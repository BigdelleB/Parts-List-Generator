import math
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
import pandas as pd


'''
Input the file path for the bom_list 
Return the categories to put into the excel file

SHOULD BE [ITEM, QTY, REFERENCE, DESCRIPTION, MANUFACTURER, MANUFACTURER P/N, MIC]
'''
def get_columns():
	columns = ["ITEM", "QTY", "REFERENCE", "DESCRIPTION", "MANUFACTURER","MANUFACTURER P/N", "MIC"]
	return columns;
  
'''
Return an array of every part in the parts list.
EVERY ROW OF PARTS HAS ["ITEM", "QTY", "REFERENCE", "DESCRIPTION", "MANUFACTURER","MANUFACTURER P/N"]
	w/o the MIC part, ignore that for now, let the User fill it in because I dont see any category for that on the parts list
'''
def get_parts_list(df):
	parts = []
	ITEM=""
	QTY=""
	REFERENCE=""
	DESCRIPTION=""
	MANUFACTURER=""
	MANUFACTURER_PN=""
	found = False

	#pandas makes reading an excel a lost easier, will return 6 columns. Need to filter out the nonsense.
	

	parts = df.loc[13:df.shape[0],:]

			
	return parts



'''
Return the number of items in the list
'''
def get_num_items(df):
	return (df.shape[0] - 13)



def main(BOM_list, file_name, cover_page):
	#define an excel file, this is the output

	df = pd.read_excel(BOM_list, usecols="A:F")
	df.columns = ["ITEM", "QTY", "REFERENCE", "DESCRIPTION","MANUFACTURER","MANUFACTURER_PN"]


	categories = get_columns()
	parts = get_parts_list(df)
	numparts = float(get_num_items(df))

	writer = pd.ExcelWriter("PL"+file_name+"-0001_.xlsx", engine='openpyxl')

	#new line
	book =  load_workbook(cover_page) 
	writer.book = book

	for i in range(0,math.ceil(numparts/27)):
		chunk = parts.loc[i*27+13:(i+1)*27+13,:]
		chunk.to_excel(writer, sheet_name="Sheet"+str(i), index= False)
		

	writer.save()




	# i=-1;

	# #make sure each parts page has max of 27 parts
	# num_pages = math.ceil(numparts/27);

	# for page in range(num_pages):
	# 	#create a new page
	# 	wb.create_sheet("Page "+ str(1+page))
	# 	ws = wb["Page "+ str(1+page)]
	# 	ws.append(categories)

	# 	if(page < num_pages-1):
	# 		for i in range(page*26,(page+1)*26):
	# 			#print(str(page)+" " +str(i))

	# 			row = parts.loc[i,:]
	# 			print(row)
	# 			ws.append(row)

	# 	else:
	# 		for i in range(page*26,len(parts)):
	# 			row = parts.loc[i,:]
	# 			ws.append(row)




	#wb.save("PL"+file_name+"-0001_.xlsx")













